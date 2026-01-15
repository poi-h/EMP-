import numpy as np
import pywt
import os
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from scipy import signal

# 数据写入表格
def write_to_excel(file_path, data, sheet_name='sheet1', col_index=1, header=None):
    """
    将数据写入 Excel 的指定列
    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param data: 一维列表或数组
    :param col_index: 目标列索引（从 1 开始）
    :param header: 列标题（可选）
    """
    # 如果文件不存在，就新建一个
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
    else:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

    # 如果有表头，写在第一行
    start_row = 1
    if header is not None:
        ws.cell(row=1, column=col_index, value=header)
        start_row = 2

    # 写入数据
    for i, value in enumerate(data, start=start_row):
        ws.cell(row=i, column=col_index, value=value)

    wb.save(file_path)

def signal_write(signal, t, a, shot_id, save_dir='.', col_index=1):  
    """
    参数/Args:
        signal: 输入信号 (Input signal)
        t: 时间序列 (Time array)
        a: 标记名 (Label for file name and title)
        save_dir: 图片保存目录 (Directory to save the image)
    返回/Returns:
        完整路径 (Full path of saved image)
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    
    fn = os.path.join(save_dir, f'signal_ch{a}.xlsx')

    if col_index == 2:
        write_to_excel(fn,  data=t, col_index=1, header='Time(s)')
    
    write_to_excel(fn,  data=signal, col_index=col_index, header=shot_id)

    return fn

def fft_write(signal, fs, a, shot_id, save_dir='.', col_index=1):
    """
    参数/Args:
        signal: 输入信号 (Input signal)
        fs: 采样率 (Sampling frequency)
        a: 标记名 (Label for file name and title)
        save_dir: 图片保存目录 (Directory to save the image)
        xlim: x轴范围 (X-axis limit for frequency plot)
    返回/Returns:
        完整路径 (Full path of saved image)
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    L = len(signal)
    Y = np.fft.fft(signal)
    P2 = np.abs(Y / L)  # 双边谱
    P1 = P2[:L // 2]  # 单边谱
    if L > 2:
        P1[1:-1] = 2 * P1[1:-1]  # 补偿能量
    f = fs * np.arange(0, L // 2) / L

    fn = os.path.join(save_dir, f'fft_ch{a}.xlsx')

    if col_index == 2:
        write_to_excel(fn,  data=f, col_index=1, header='Frequency(Hz)')
    
    write_to_excel(fn,  data=P1, col_index=col_index, header=shot_id)
    
    return fn

def fft_plot_Ae(signal, fs, a, shot_id, save_dir='.', G_dBi=4, c=3e8, col_index=1):
    """
    对信号 signal 做FFT分析，使用有效面积A_e修正并保存频谱图。
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    L = len(signal)
    Y = np.fft.fft(signal)
    P2 = np.abs(Y / L)  # 双边谱
    P1 = P2[:L // 2]    # 单边谱
    if L > 2:
        P1[1:-1] = 2 * P1[1:-1]  # 补偿能量
    f = fs * np.arange(0, L // 2) / L

    # 有效面积修正
    G_linear = 10 ** (G_dBi / 10)
    wavelengths = c / f
    A_e = G_linear * wavelengths ** 2 / (4 * np.pi)
    # 避免除以零
    # A_e[P1 == 0] = np.nan
    P1_Ae = P1 / A_e

    if L > 2:
        P1[1:-1] = 2 * P1[1:-1]  # 补偿能量
    f = fs * np.arange(0, L // 2) / L

    fn = os.path.join(save_dir, f'fft_ch{a}.xlsx')

    if col_index == 2:
        write_to_excel(fn,  data=f, col_index=1, header='Frequency(Hz)')
    
    write_to_excel(fn,  data=P1, col_index=col_index, header=shot_id)

    return fn

def cwt_plot(signal, t, fs, a, shot_id, save_dir='.', totalscal=8192, wavename='cmor1-1', xlim=(0, 6e-8), ylim=(0, 6e9)):
    """
    对信号 signal 做连续小波变换并保存图像。
    Perform continuous wavelet transform (CWT) on the signal and save the image.

    参数/Args:
        signal: 输入信号 (Input signal)
        t: 时间序列 (Time array)
        dt: 时间分辨率 (Time step)
        a: 标记名 (Label for file name and title)
        save_dir: 图片保存目录 (Directory to save the image)
        totalscal: 小波尺度总数 (Total number of scales, default 8192)
        wavename: 小波名称 (Wavelet name, default 'cmor1-1')
    返回/Returns:
        图片完整路径 (Full path of saved image)
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    Fc = pywt.central_frequency(wavename)
    c = 2 * Fc * totalscal
    scals = c / np.arange(1, totalscal + 1)
    f = fs * pywt.scale2frequency(wavename, scals)
    coefs, freqs = pywt.cwt(signal, scals, wavename)
    plt.figure()
    plt.imshow(np.abs(coefs), aspect='auto', extent=[t.min(), t.max(), f.max(), f.min()], cmap='jet')
    plt.colorbar()
    plt.title(f'Continuous Wavelet Transform {a}')
    plt.xlabel('t(s)')
    plt.ylabel('Frequency (Hz)')
    plt.gca().xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='x', scilimits=(0,0))
    plt.gca().yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='y', scilimits=(0,0))
    plt.xlim(xlim)
    plt.ylim(ylim)
    fn = os.path.join(save_dir, f'{shot_id:03d}wavelet{a}.png')
    plt.savefig(fn, dpi=600)
    plt.close()
    return fn

# 绘图
def signal_plot(signal, t, a, shot_id, save_dir='.', xlim=(0, 6e-8), ylim=(-1e5, 1e5)):  
    """
    绘制时域信号并保存为图片。
    Plot the time-domain signal and save as an image.
    
    参数/Args:
        signal: 输入信号 (Input signal)
        t: 时间序列 (Time array)
        a: 标记名 (Label for file name and title)
        save_dir: 图片保存目录 (Directory to save the image)
    返回/Returns:
        图片完整路径 (Full path of saved image)
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    plt.figure()
    plt.plot(t, signal)
    plt.title(f'Electric Field {a}')
    plt.xlabel('t(s)')
    plt.ylabel('E(V/m)')
    plt.gca().xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='x', scilimits=(0,0))
    plt.gca().yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='y', scilimits=(0,0))
    plt.xlim(xlim)
    plt.ylim(ylim)
    fn = os.path.join(save_dir, f'{shot_id:03d}E{a}.png')
    plt.savefig(fn, dpi=600)
    plt.close()
    return fn

def fft_plot(signal, fs, a, shot_id, save_dir='.', xlim=(0, 6e9)):
    """
    对信号 signal 做FFT分析并保存频谱图。
    Perform FFT analysis on the signal and save the spectrum image.

    参数/Args:
        signal: 输入信号 (Input signal)
        fs: 采样率 (Sampling frequency)
        a: 标记名 (Label for file name and title)
        save_dir: 图片保存目录 (Directory to save the image)
        xlim: x轴范围 (X-axis limit for frequency plot)
    返回/Returns:
        图片完整路径 (Full path of saved image)
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    L = len(signal)
    Y = np.fft.fft(signal)
    P2 = np.abs(Y / L)  # 双边谱
    P1 = P2[:L // 2]  # 单边谱
    if L > 2:
        P1[1:-1] = 2 * P1[1:-1]  # 补偿能量
    f = fs * np.arange(0, L // 2) / L
    plt.figure()
    plt.plot(f, P1, color='r')
    plt.title(f'Single-Sided Amplitude Spectrum {a}')
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('|P1|')
    plt.gca().xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='x', scilimits=(0,0))
    plt.gca().yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='y', scilimits=(0,0))
    plt.xlim(xlim)
    fn = os.path.join(save_dir, f'{shot_id:03d}fft{a}.png')
    plt.savefig(fn, dpi=600)
    plt.close()
    return fn

def fft_plot_Ae(signal, fs, a, shot_id, save_dir='.', xlim=(0, 6e9), G_dBi=4, c=3e8):
    """
    对信号 signal 做FFT分析，使用有效面积A_e修正并保存频谱图。
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    L = len(signal)
    Y = np.fft.fft(signal)
    P2 = np.abs(Y / L)  # 双边谱
    P1 = P2[:L // 2]    # 单边谱
    if L > 2:
        P1[1:-1] = 2 * P1[1:-1]  # 补偿能量
    f = fs * np.arange(0, L // 2) / L

    # 有效面积修正
    G_linear = 10 ** (G_dBi / 10)
    wavelengths = c / f
    A_e = G_linear * wavelengths ** 2 / (4 * np.pi)
    # 避免除以零
    # A_e[P1 == 0] = np.nan
    P1_Ae = P1 / A_e

    plt.figure()
    plt.plot(f, P1_Ae, color='g')
    plt.title(f' shot 044 FFT(A_e corrected) {a}')
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('|P1| / A_e')
    plt.gca().xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='x', scilimits=(0,0))
    plt.gca().yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='y', scilimits=(0,0))
    plt.xlim(xlim)
    fn = os.path.join(save_dir, f'{shot_id:03d}fft_Ae{a}.png')
    plt.savefig(fn, dpi=600)
    plt.close()
    return fn

def cwt_plot(signal, t, fs, a, shot_id, save_dir='.', totalscal=8192, wavename='cmor1-1', xlim=(0, 6e-8), ylim=(0, 6e9)):
    """
    对信号 signal 做连续小波变换并保存图像。
    Perform continuous wavelet transform (CWT) on the signal and save the image.

    参数/Args:
        signal: 输入信号 (Input signal)
        t: 时间序列 (Time array)
        dt: 时间分辨率 (Time step)
        a: 标记名 (Label for file name and title)
        save_dir: 图片保存目录 (Directory to save the image)
        totalscal: 小波尺度总数 (Total number of scales, default 8192)
        wavename: 小波名称 (Wavelet name, default 'cmor1-1')
    返回/Returns:
        图片完整路径 (Full path of saved image)
    """
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    Fc = pywt.central_frequency(wavename)
    c = 2 * Fc * totalscal
    scals = c / np.arange(1, totalscal + 1)
    f = fs * pywt.scale2frequency(wavename, scals)
    coefs, freqs = pywt.cwt(signal, scals, wavename)
    plt.figure()
    plt.imshow(np.abs(coefs), aspect='auto', extent=[t.min(), t.max(), f.max(), f.min()], cmap='jet')
    plt.colorbar()
    plt.title(f'Continuous Wavelet Transform {a}')
    plt.xlabel('t(s)')
    plt.ylabel('Frequency (Hz)')
    plt.gca().xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='x', scilimits=(0,0))
    plt.gca().yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
    plt.ticklabel_format(style='sci', axis='y', scilimits=(0,0))
    plt.xlim(xlim)
    plt.ylim(ylim)
    fn = os.path.join(save_dir, f'{shot_id:03d}wavelet{a}.png')
    plt.savefig(fn, dpi=600)
    plt.close()
    return fn

# 高通滤波
def highpass_filter(signal, fs, cutoff_freq):
    nyquist = 0.5 * fs
    normal_cutoff = cutoff_freq / nyquist
    b, a = signal.butter(4, normal_cutoff, btype='high')
    filtered_signal = signal.filtfilt(b, a, signal)
    return filtered_signal

# 低通滤波器
def lowpass_filter(signal, fs, cutoff_freq):
    nyquist = 0.5 * fs
    normal_cutoff = cutoff_freq / nyquist
    b, a = signal.butter(4, normal_cutoff, btype='low')
    filtered_signal = signal.filtfilt(b, a, signal)
    return filtered_signal

# B-dot信号积分
def process_signal(t, signal, gain = 1, A = 1):
    dt = np.diff(t)
    B = np.zeros_like(signal)
    for i in range(1,len(signal)):
        B[i] = B[i-1] + (signal[i]/gain)/A * dt[i-1]
    return B

# 积分能量
def integrate_energy(t, signal):
    dt = np.diff(t)
    energy = 0.0
    for i in range(1, len(signal)):
        energy += 0.5 * (signal[i]**2 + signal[i-1]**2) * dt[i-1]
    return energy

# 求信号包络线
def envelope(signal):
    analytic_signal = signal.hilbert(signal)
    envelope = np.abs(analytic_signal)
    return envelope

# 绘图并标记峰值
def plot_with_market(t, signal, x0, y0, shot_id, a, save_dir='.'):
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    plt.figure()
    plt.plot(t, signal, label='Signal')
    plt.annotate("",
             xy=(x0, y0),
             xytext=(x0+3e-9, y0*1.1),
             arrowprops=dict(arrowstyle="->"))
    plt.title(f'Signal with Peaks {a}')
    plt.xlabel('t(s)')
    plt.ylabel('Signal')
    plt.legend()
    fn = os.path.join(save_dir, f'{shot_id:03d}peaks{a}.png')
    plt.savefig(fn, dpi=600)
    plt.close()
    return fn